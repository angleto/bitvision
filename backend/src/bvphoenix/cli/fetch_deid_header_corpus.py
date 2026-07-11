"""CLI: fetch the TCIA Pseudo-PHI-DICOM-Data corpus for the header-engine gate.

Downloads the collection's DICOM (NBIA v1, reusing the hardened TCIA adapter of
``bvphoenix-public-import``) into a LOCAL corpus dir and converts the
collection's ground-truth answer key (a spreadsheet distributed on the TCIA
collection page — download it manually with a browser, its URL is not stable)
into the ``answer_key_header.json`` consumed by
``services.header_deid_eval.load_header_corpus``.

This corpus must NEVER enter the OpenData library (it deliberately contains
pseudo-PHI) — that is why it does not go through ``bvphoenix-public-import``
and why the manifest keeps the collection commented out. Local dir + optional
private dataset-bucket sync only.

    bvphoenix-fetch-deid-header-corpus --out ~/data/deid-header-corpus \
        --subjects 3                       # smoke slice; omit for all
    bvphoenix-fetch-deid-header-corpus --out ... --answer-key ~/Downloads/key.csv
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import click
import httpx

from bvphoenix.cli._public_http import HTTP_TIMEOUT
from bvphoenix.cli.public_import import _adapter_tcia, _adapter_tcia_list_patients
from bvphoenix.services.header_deid_eval import ANSWER_KEY_NAME

COLLECTION = "Pseudo-PHI-DICOM-Data"

_SOP_COL_RE = re.compile(r"sop.*(uid|instance)", re.IGNORECASE)


def _read_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    """Read the answer-key spreadsheet as (header, rows). CSV/TSV via stdlib;
    XLSX via openpyxl when installed (dev extra). The exact format TCIA ships
    is not pinned — the parser is deliberately tolerant and the caller sees a
    clear error when the SOP-UID column cannot be located."""
    if path.suffix.lower() in (".csv", ".tsv", ".txt"):
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            rows = [row for row in reader if any(c.strip() for c in row)]
    elif path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dev-extra guidance
            raise click.ClickException(
                "reading .xlsx needs openpyxl (uv sync --extra dev)"
            ) from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [
            ["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() for c in row)
        ]
    else:
        raise click.ClickException(f"unsupported answer-key format: {path.suffix}")
    if not rows:
        raise click.ClickException("answer key is empty")
    return [h.strip() for h in rows[0]], [list(r) for r in rows[1:]]


def derive_answer_key(root: Path) -> int:
    """Build answer_key_header.json from the corpus instances' own headers
    (see ``header_deid_eval.derive_case_phi``)."""
    import io

    import pydicom

    from bvphoenix.services.header_deid_eval import derive_case_phi

    key: dict[str, list[dict]] = {}
    for path in sorted(root.rglob("*.dcm")):
        try:
            ds = pydicom.dcmread(io.BytesIO(path.read_bytes()), stop_before_pixels=True)
        except Exception:
            continue
        sop_uid = str(getattr(ds, "SOPInstanceUID", "") or "")
        if not sop_uid:
            continue
        values = [{"value": p.value, "category": p.category} for p in derive_case_phi(ds)]
        if values:
            key[sop_uid] = values
    (root / ANSWER_KEY_NAME).write_text(json.dumps(key, indent=2, ensure_ascii=False))
    return len(key)


def convert_answer_key(src: Path, out_dir: Path) -> int:
    """Normalize the TCIA answer-key spreadsheet to answer_key_header.json.

    Heuristic mapping: the column whose header matches SOP-UID keys the
    entries; every other non-empty cell becomes one PHI value whose category
    is the (slugified) column header. Values stay human-auditable — the eval
    normalizes at comparison time, not here."""
    header, rows = _read_rows(src)
    sop_idx = next((i for i, h in enumerate(header) if _SOP_COL_RE.search(h)), None)
    if sop_idx is None:
        raise click.ClickException(
            f"could not locate a SOPInstanceUID column in {header!r} — "
            "inspect the spreadsheet and rename the column to contain 'SOP...UID'"
        )
    key: dict[str, list[dict]] = {}
    for row in rows:
        if sop_idx >= len(row):
            continue
        sop_uid = str(row[sop_idx]).strip()
        if not sop_uid:
            continue
        values = []
        for i, cell in enumerate(row):
            if i == sop_idx:
                continue
            text = str(cell).strip()
            if not text:
                continue
            category = re.sub(r"[^a-z0-9]+", "_", header[i].lower()).strip("_") or "other"
            values.append({"value": text, "category": category})
        if values:
            key.setdefault(sop_uid, []).extend(values)
    (out_dir / ANSWER_KEY_NAME).write_text(json.dumps(key, indent=2, ensure_ascii=False))
    return len(key)


@click.command()
@click.option("--out", "out_dir", required=True, type=click.Path(), help="Corpus directory.")
@click.option(
    "--subjects",
    default=0,
    show_default=True,
    help="Limit to the first N subjects (0 = the whole collection).",
)
@click.option(
    "--answer-key",
    "answer_key",
    type=click.Path(exists=True, path_type=Path),
    default=None,
    help="Answer-key spreadsheet (csv/tsv/xlsx) downloaded from the TCIA page.",
)
@click.option(
    "--skip-download",
    is_flag=True,
    help="Only (re)build the answer key over an existing corpus dir.",
)
def main(out_dir: str, subjects: int, answer_key: Path | None, skip_download: bool) -> None:
    root = Path(out_dir)
    root.mkdir(parents=True, exist_ok=True)

    if not skip_download:
        with httpx.Client(timeout=HTTP_TIMEOUT) as client:
            ids = _adapter_tcia_list_patients(client, collection=COLLECTION)
            if subjects:
                ids = ids[:subjects]
            click.echo(f"{COLLECTION}: fetching {len(ids)} subject(s) into {root}")
            for n, subject_id in enumerate(ids, start=1):
                click.echo(f"  [{n}/{len(ids)}] {subject_id}")
                _adapter_tcia(client, collection=COLLECTION, subject_id=subject_id, workdir=root)

    if answer_key is not None:
        count = convert_answer_key(answer_key, root)
        click.echo(f"answer key (spreadsheet): {count} entr(ies) -> {root / ANSWER_KEY_NAME}")
    else:
        # TCIA distributes UID crosswalks, not a per-value key: the planted
        # synthetic PHI IS the content of the identifying header attributes,
        # so the ground truth is derived from the corpus itself (auditable
        # JSON, reviewed like any other fixture).
        count = derive_answer_key(root)
        click.echo(f"answer key (derived from headers): {count} entr(ies)")
    click.echo(f"done. Point BVP_DEID_HEADER_CORPUS={root} and run the gate:")
    click.echo("  uv run pytest tests/test_deid_header_corpus.py -q")


if __name__ == "__main__":
    main()
